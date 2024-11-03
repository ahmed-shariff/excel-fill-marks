import openpyxl
import easygui
import sys


def _main():
    file_name = easygui.fileopenbox(msg="Select file", title="Marking file selection", filetypes="*.xlsx")
    # file_name = "C:/Users/amsha/Projects/excel-fill-marks/GROUP NO_ELA2204 FINAL MARKS_IN CHARGE.xlsx"
    workbook = openpyxl.load_workbook(file_name)
    if len(workbook.sheetnames) > 1:
        sheet_to_use = easygui.choicebox(msg=f"Loading workbook {file_name}. Which sheet to use?", choices=workbook.sheetnames)
    else:
        sheet_to_use = workbook.sheetnames[0]

    sheet = workbook[sheet_to_use]
    reg_first_cell_id = easygui.enterbox(msg="Provide first cell with registration number (e.g., D5):", strip=True)
    # reg_first_cell_id = "d5"

    reg_first_cell = sheet[reg_first_cell_id]

    valid_reg_numbers = {}
    for (_cell, ) in sheet.iter_rows(min_col=reg_first_cell.column, max_col=reg_first_cell.column, min_row=reg_first_cell.row):
        if _cell.value is None:
            break

        if _cell.value in valid_reg_numbers:
            easygui.msgbox("Dupliate Reg IDs")
            sys.exit()
        valid_reg_numbers[_cell.value] = _cell

    success = easygui.textbox(msg="Registration numbers found (editing will not make a difference here!). Click OK to proceed. Clicking Cancel will quit the application.", text=[val + "\n" for val in valid_reg_numbers.keys()])

    if success is None:
        sys.exit()

    while True:
        marks_column_id = easygui.enterbox(msg="Provide column where marks are to be entered (e.g., F):", strip=True)
        if marks_column_id is None and easygui.ynbox("exit?"):
            sys.exit()

        if not marks_column_id.isalpha() or len(marks_column_id) != 1:
            easygui.msgbox(f"Makrs column id is not valid (got {marks_column_id})")
            continue

        if sheet[marks_column_id][-1].column == reg_first_cell.column:
            easygui.msgbox("Makrs column id is the same as the reg column!")
            continue

        break

    easygui.msgbox(msg=f"Doing the following:\n Selected file: {file_name}\n Selected sheet: {sheet_to_use}\n Registration ids starting from cell: {reg_first_cell_id}\n Entering marks in column: {marks_column_id}")

    maybe_close = False
    error_msg = ""
    while True:
        if maybe_close:
            maybe_close = False
            if easygui.ynbox("Close application?"):
                while True:
                    save_location = easygui.filesavebox("Save workbook to:")
                    if save_location is None:
                        if easygui.ynbox("Do you want to close without saving? Click cancel if you don't want to save the file. Click ok if you want to try save the file again."):
                            continue
                        else:
                            break
                    else:
                        workbook.save(save_location)
                        break
                sys.exit()

        values = easygui.multenterbox(msg=f"{error_msg}Enter reg number and marks", fields=["reg number", "marks"])
        error_msg = ""

        if values:
            _reg_num, _mark = values
            for _id, _cell in valid_reg_numbers.items():
                if _reg_num in _id:
                    try:
                        _mark = float(_mark)
                        _mark_cell = f"{marks_column_id}{_cell.row}"
                        sheet[_mark_cell] = _mark
                        break
                    except ValueError:
                        error_msg = f"Invalid marks (got {_mark})\n\n"
                        break
            else:
                error_msg = f"Invalid reg number (got {_reg_num})\n\n"
        else:
            maybe_close = True


if __name__ == "__main__":
    _main()
